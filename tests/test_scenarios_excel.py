from datetime import date

import pandas as pd
from openpyxl import load_workbook

from agent_trading_signal.backtest.scenarios import ScenarioDefinition, run_scenarios
from agent_trading_signal.reporting.excel import write_scenario_workbook
from agent_trading_signal.settings import (
    AssetConfig,
    BacktestConfig,
    SignalConfig,
    UniverseConfig,
)


def test_write_scenario_workbook(tmp_path) -> None:
    length = 520
    prices = pd.DataFrame(
        {
            "A": [100 * (1.003**index) for index in range(length)],
            "B": [100 * (1.001**index) for index in range(length)],
            "C": [100 * (0.999**index) for index in range(length)],
        },
        index=pd.date_range(start="2019-01-01", periods=length, freq="D"),
    )
    universe = UniverseConfig(
        assets=[
            AssetConfig(
                symbol="A",
                name="Asset A",
                asset_class="test",
                price_symbol="A",
                trade_symbol="A",
            ),
            AssetConfig(
                symbol="B",
                name="Asset B",
                asset_class="test",
                price_symbol="B",
                trade_symbol="B",
            ),
            AssetConfig(
                symbol="C",
                name="Asset C",
                asset_class="test",
                price_symbol="C",
                trade_symbol="C",
            ),
        ]
    )
    scenarios = run_scenarios(
        prices=prices,
        universe=universe,
        signal_config=SignalConfig(),
        backtest_config=BacktestConfig(start=date(2020, 1, 1), benchmarks=["A"]),
        scenario_definitions=[
            ScenarioDefinition("Baseline", require_above_sma200_for_entries=False),
            ScenarioDefinition("SMA200 filter", require_above_sma200_for_entries=True),
        ],
    )
    output = tmp_path / "analysis.xlsx"

    write_scenario_workbook(scenarios, output)

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Dashboard",
        "Trades",
        "Worst Trades",
        "Asset PnL",
        "Equity Curves",
        "Drawdowns",
    ]
    assert len(workbook["Dashboard"]._charts) == 2
    assert "Outcome Band" in [cell.value for cell in workbook["Trades"][3]]


def test_run_scenarios_skips_incompatible_exclusions() -> None:
    length = 520
    prices = pd.DataFrame(
        {
            "A": [100 * (1.003**index) for index in range(length)],
            "B": [100 * (1.001**index) for index in range(length)],
        },
        index=pd.date_range(start="2019-01-01", periods=length, freq="D"),
    )
    universe = UniverseConfig(
        assets=[
            AssetConfig(
                symbol="A",
                name="Asset A",
                asset_class="test",
                price_symbol="A",
                trade_symbol="A",
            ),
            AssetConfig(
                symbol="B",
                name="Asset B",
                asset_class="test",
                price_symbol="B",
                trade_symbol="B",
            ),
        ]
    )

    scenarios = run_scenarios(
        prices=prices,
        universe=universe,
        signal_config=SignalConfig(),
        backtest_config=BacktestConfig(start=date(2020, 1, 1), benchmarks=["A"]),
        scenario_definitions=[
            ScenarioDefinition("A/B", require_above_sma200_for_entries=True),
            ScenarioDefinition(
                "Invalid ex ETH",
                require_above_sma200_for_entries=True,
                excluded_symbols=("ETH",),
            ),
        ],
    )

    assert [scenario.definition.name for scenario in scenarios] == ["A/B"]
