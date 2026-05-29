from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from agent_trading_signal.backtest.scenarios import (
    ScenarioBacktest,
    asset_contribution_frame,
    drawdown_frame,
    monthly_equity_frame,
    scenario_summary_frame,
    trade_analysis_frame,
    worst_trades_frame,
)

MONEY_FORMAT = "$#,##0.00;[Red]($#,##0.00);-"
PERCENT_FORMAT = "0.00%;[Red](0.00%);-"
NUMBER_FORMAT = "#,##0.00"
DATE_FORMAT = "yyyy-mm-dd"
TITLE_FILL = "17324D"
HEADER_FILL = "D9EAF7"
NOTE_FILL = "F4F7FA"
GAIN_FILL = "D9EAD3"
NEUTRAL_FILL = "FFF2CC"
LOSS_FILL = "F4CCCC"


def write_scenario_workbook(scenarios: list[ScenarioBacktest], path: str | Path) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    trades = workbook.create_sheet("Trades")
    worst = workbook.create_sheet("Worst Trades")
    assets = workbook.create_sheet("Asset PnL")
    curves = workbook.create_sheet("Equity Curves")
    drawdowns = workbook.create_sheet("Drawdowns")

    summary = scenario_summary_frame(scenarios)
    trade_analysis = trade_analysis_frame(scenarios)
    worst_trades = worst_trades_frame(scenarios)
    asset_contribution = asset_contribution_frame(scenarios)
    equity = monthly_equity_frame(scenarios)
    scenario_names = [scenario.definition.name for scenario in scenarios]
    drawdown = drawdown_frame(equity, scenario_names)

    _write_dashboard(dashboard, summary, equity, drawdown)
    _write_table_sheet(trades, "All Trades with Realized PnL", trade_analysis, "Trades")
    _write_table_sheet(worst, "Worst Trades by Return", worst_trades, "WorstTrades")
    _write_table_sheet(assets, "Approximate Asset Contribution", asset_contribution, "AssetPnL")
    _write_table_sheet(curves, "Monthly Equity Curves", equity, "EquityCurves")
    _write_table_sheet(drawdowns, "Monthly Drawdowns", drawdown, "Drawdowns")
    _add_dashboard_charts(dashboard, curves, drawdowns, equity, drawdown)

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    workbook.save(output_path)
    _verify_workbook(output_path)


def _write_dashboard(
    sheet,
    summary: pd.DataFrame,
    equity: pd.DataFrame,
    drawdown: pd.DataFrame,
) -> None:
    _title(sheet, "A1:M1", "Agent Trading Signal - Strategy Analysis")
    _write_frame(sheet, summary, start_row=3, start_col=1)
    _add_table(sheet, 3, 1, len(summary) + 1, len(summary.columns), "ScenarioSummary")
    _format_columns(sheet, summary, header_row=3)
    sheet.freeze_panes = "A4"

    note = (
        "Interpretation: the SMA200 entry filter improves both return and drawdown versus "
        "the baseline. Excluding ETH and SLV improves this sample further, mostly by avoiding "
        "the largest ETH/SLV drawdown events. The Trades tab color-codes realized outcomes and "
        "conviction. This is research output, not a finalized production rule."
    )
    sheet.merge_cells("A9:M12")
    cell = sheet["A9"]
    cell.value = note
    cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(9, 13):
        sheet.row_dimensions[row].height = 23

    widths = {
        "A": 20,
        "B": 12,
        "C": 12,
        "D": 16,
        "E": 16,
        "F": 14,
        "G": 12,
        "H": 15,
        "I": 13,
        "J": 10,
        "K": 16,
        "L": 13,
        "M": 15,
    }
    _set_widths(sheet, widths)

    if not equity.empty and not drawdown.empty:
        sheet["A14"] = "Charts use monthly observations to keep the workbook responsive."
        sheet["A14"].font = Font(italic=True, color="555555")


def _write_table_sheet(sheet, title: str, frame: pd.DataFrame, table_name: str) -> None:
    _title(sheet, f"A1:{_column_letter(max(len(frame.columns), 1))}1", title)
    if frame.empty:
        sheet["A3"] = "No data"
        return

    _write_frame(sheet, frame, start_row=3, start_col=1)
    _add_table(sheet, 3, 1, len(frame) + 1, len(frame.columns), table_name)
    _format_columns(sheet, frame, header_row=3)
    _apply_trade_coloring(sheet, frame, header_row=3)
    sheet.freeze_panes = "A4"
    _autowidth(sheet, frame)


def _write_frame(sheet, frame: pd.DataFrame, start_row: int, start_col: int) -> None:
    for col_offset, column in enumerate(frame.columns):
        sheet.cell(start_row, start_col + col_offset, column)
    for row_offset, row in enumerate(frame.itertuples(index=False), start=1):
        for col_offset, value in enumerate(row):
            cell = sheet.cell(start_row + row_offset, start_col + col_offset)
            if isinstance(value, pd.Timestamp):
                cell.value = value.to_pydatetime()
            else:
                cell.value = value


def _title(sheet, range_ref: str, title: str) -> None:
    sheet.merge_cells(range_ref)
    cell = sheet[range_ref.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[cell.row].height = 26


def _add_table(sheet, start_row: int, start_col: int, rows: int, cols: int, name: str) -> None:
    end_row = start_row + rows - 1
    end_col = start_col + cols - 1
    ref = f"{_cell_ref(start_row, start_col)}:{_cell_ref(end_row, end_col)}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for cell in sheet[start_row]:
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)


def _format_columns(sheet, frame: pd.DataFrame, header_row: int) -> None:
    for index, column in enumerate(frame.columns, start=1):
        letter = _column_letter(index)
        if column in {"Start", "End", "Signal Date", "Entry Date", "Exit Date"}:
            number_format = DATE_FORMAT
        elif column in {
            "Start Value",
            "End Value",
            "Total Cost",
            "Entry Capital",
            "Exit Capital Before Next Cost",
            "PnL",
            "Entry Cost",
            "Weighted PnL",
        }:
            number_format = MONEY_FORMAT
        elif column in {
            "Total Return",
            "CAGR",
            "Max Drawdown",
            "Volatility",
            "Average Turnover",
            "Time in Cash",
            "Return",
        }:
            number_format = PERCENT_FORMAT
        elif pd.api.types.is_numeric_dtype(frame[column]):
            number_format = NUMBER_FORMAT
        else:
            number_format = None

        if number_format:
            for cell in sheet[f"{letter}{header_row + 1}:{letter}{header_row + len(frame)}"]:
                cell[0].number_format = number_format


def _apply_trade_coloring(sheet, frame: pd.DataFrame, header_row: int) -> None:
    if frame.empty or "Return" not in frame.columns:
        return

    columns = {column: index for index, column in enumerate(frame.columns, start=1)}
    outcome_columns = [
        columns[column] for column in ("PnL", "Return", "Outcome Band") if column in columns
    ]
    conviction_column = columns.get("Conviction")

    for row in range(header_row + 1, header_row + len(frame) + 1):
        return_value = sheet.cell(row, columns["Return"]).value
        outcome_fill = _outcome_fill(return_value)
        for column in outcome_columns:
            sheet.cell(row, column).fill = outcome_fill

        if conviction_column is None:
            continue
        conviction = str(sheet.cell(row, conviction_column).value).lower()
        sheet.cell(row, conviction_column).fill = _conviction_fill(conviction)


def _outcome_fill(return_value: object) -> PatternFill:
    if not isinstance(return_value, int | float):
        return PatternFill("solid", fgColor=NEUTRAL_FILL)
    if return_value < -0.10:
        return PatternFill("solid", fgColor=LOSS_FILL)
    if return_value > 0.10:
        return PatternFill("solid", fgColor=GAIN_FILL)
    return PatternFill("solid", fgColor=NEUTRAL_FILL)


def _conviction_fill(conviction: str) -> PatternFill:
    if conviction in {"high", "strong"}:
        return PatternFill("solid", fgColor=GAIN_FILL)
    if conviction == "medium":
        return PatternFill("solid", fgColor=NEUTRAL_FILL)
    return PatternFill("solid", fgColor=LOSS_FILL)


def _add_dashboard_charts(
    dashboard,
    curves,
    drawdowns,
    equity: pd.DataFrame,
    drawdown: pd.DataFrame,
) -> None:
    if equity.empty or drawdown.empty:
        return

    equity_chart = LineChart()
    equity_chart.title = "Equity Curves"
    equity_chart.y_axis.numFmt = "$#,##0"
    equity_chart.y_axis.title = "Portfolio value"
    equity_chart.x_axis.title = "Month"
    equity_data = Reference(
        curves, min_col=2, max_col=equity.shape[1], min_row=3, max_row=3 + len(equity)
    )
    equity_categories = Reference(curves, min_col=1, min_row=4, max_row=3 + len(equity))
    equity_chart.add_data(equity_data, titles_from_data=True)
    equity_chart.set_categories(equity_categories)
    equity_chart.height = 9
    equity_chart.width = 17
    dashboard.add_chart(equity_chart, "O3")

    drawdown_chart = LineChart()
    drawdown_chart.title = "Strategy Drawdowns"
    drawdown_chart.y_axis.numFmt = "0%"
    drawdown_chart.y_axis.title = "Drawdown"
    drawdown_chart.x_axis.title = "Month"
    dd_data = Reference(
        drawdowns, min_col=2, max_col=drawdown.shape[1], min_row=3, max_row=3 + len(drawdown)
    )
    dd_categories = Reference(drawdowns, min_col=1, min_row=4, max_row=3 + len(drawdown))
    drawdown_chart.add_data(dd_data, titles_from_data=True)
    drawdown_chart.set_categories(dd_categories)
    drawdown_chart.height = 9
    drawdown_chart.width = 17
    dashboard.add_chart(drawdown_chart, "O22")


def _autowidth(sheet, frame: pd.DataFrame) -> None:
    for index, column in enumerate(frame.columns, start=1):
        max_length = len(str(column))
        sample = frame[column].head(100)
        if not sample.empty:
            max_length = max(max_length, *(len(str(value)) for value in sample))
        sheet.column_dimensions[_column_letter(index)].width = min(max(max_length + 2, 10), 32)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    required = {"Dashboard", "Trades", "Worst Trades", "Asset PnL", "Equity Curves", "Drawdowns"}
    missing = required - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Missing workbook sheets: {', '.join(sorted(missing))}")
    if workbook["Dashboard"].max_row < 10:
        raise ValueError("Dashboard sheet appears incomplete")


def _cell_ref(row: int, column: int) -> str:
    return f"{_column_letter(column)}{row}"


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
